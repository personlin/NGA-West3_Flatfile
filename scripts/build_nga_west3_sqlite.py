#!/usr/bin/env python3
"""
Build a normalized SQLite database from the NGA-West3 public flatfiles.

The public release contains flatfiles rather than the complete operational
GMDB. This script reconstructs a normalized, application-friendly schema from
the release files while preserving every ground-motion ordinate in the CSVs.
"""

from __future__ import annotations

import csv
import hashlib
import json
import re
import sqlite3
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "output" / "sqlite" / "nga_west3_20250919.sqlite"

SOURCE_CSV = ROOT / "NGA_West3_Source_Metadata_20250919.csv"
STATION_CSV = ROOT / "NGA_West3_Station_Metadata_20250919.csv"
DOC_XLSX = ROOT / "NGA_West3_Flatfile_Documentation_20250919.xlsx"
C1C2_XLSX = ROOT / "NGA_West3_C1C2_North_America_20250919.xlsx"

COMPONENT_FILES = {
    "H1": ROOT / "NGA_West3_H1_Flatfile_20250919.csv",
    "H2": ROOT / "NGA_West3_H2_Flatfile_20250919.csv",
    "V": ROOT / "NGA_West3_V_Flatfile_20250919.csv",
    "RotD0": ROOT / "NGA_West3_RotD0_Flatfile_20250919.csv",
    "RotD50": ROOT / "NGA_West3_RotD50_Flatfile_20250919.csv",
    "RotD100": ROOT / "NGA_West3_RotD100_Flatfile_20250919.csv",
}
EAS_CSV = ROOT / "NGA_West3_EAS_Flatfile_20250919.csv"

MOTION_BASE_COMPONENT = "H1"
BATCH_SIZE = 1000
MISSING = {"", "NA", "NaN", "nan", "NULL"}


def q(name: str) -> str:
    return '"' + name.replace('"', '""') + '"'


def scalar(value: str | None):
    if value is None:
        return None
    if not isinstance(value, str):
        return value
    value = value.strip()
    if value in MISSING:
        return None
    if re.fullmatch(r"[-+]?\d+", value):
        try:
            return int(value)
        except ValueError:
            return value
    if re.fullmatch(r"[-+]?(?:\d+\.\d*|\.\d+|\d+)(?:[eE][-+]?\d+)?", value):
        try:
            return float(value)
        except ValueError:
            return value
    return value


def float_or_none(value: str | None):
    value = scalar(value)
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def has_real_value(values) -> bool:
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if text and text not in {"-999", "-999.0", "-999.00"}:
            return True
    return False


def parse_axis_value(token: str) -> float:
    return float(token.replace("p", "."))


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def csv_header(path: Path) -> list[str]:
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as fh:
        return next(csv.reader(fh))


def create_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        PRAGMA journal_mode = WAL;
        PRAGMA synchronous = OFF;
        PRAGMA temp_store = MEMORY;
        PRAGMA foreign_keys = OFF;

        DROP TABLE IF EXISTS release_files;
        DROP TABLE IF EXISTS field_catalog;
        DROP TABLE IF EXISTS code_definitions;
        DROP TABLE IF EXISTS citations;
        DROP TABLE IF EXISTS event_types;
        DROP TABLE IF EXISTS events;
        DROP TABLE IF EXISTS finite_faults;
        DROP TABLE IF EXISTS finite_fault_kinematic_parameters;
        DROP TABLE IF EXISTS finite_fault_segments;
        DROP TABLE IF EXISTS networks;
        DROP TABLE IF EXISTS stations;
        DROP TABLE IF EXISTS sites;
        DROP TABLE IF EXISTS basin_depth_estimates;
        DROP TABLE IF EXISTS motions;
        DROP TABLE IF EXISTS paths;
        DROP TABLE IF EXISTS time_series_metadata;
        DROP TABLE IF EXISTS intensity_measures;
        DROP TABLE IF EXISTS response_spectra;
        DROP TABLE IF EXISTS effective_amplitude_spectra;
        DROP TABLE IF EXISTS c1c2_classifications;
        DROP TABLE IF EXISTS spectral_axes;

        CREATE TABLE release_files (
          file_name TEXT PRIMARY KEY,
          file_role TEXT NOT NULL,
          component TEXT,
          row_count INTEGER,
          column_count INTEGER,
          sha256 TEXT
        );

        CREATE TABLE field_catalog (
          source_sheet TEXT NOT NULL,
          column_label TEXT,
          field_name TEXT NOT NULL,
          data_type TEXT,
          description TEXT,
          PRIMARY KEY (source_sheet, field_name)
        );

        CREATE TABLE code_definitions (
          category TEXT NOT NULL,
          code TEXT NOT NULL,
          label TEXT,
          description TEXT,
          raw_json TEXT,
          PRIMARY KEY (category, code)
        );

        CREATE TABLE citations (
          citation_id INTEGER PRIMARY KEY,
          citation TEXT,
          doi_or_url TEXT
        );

        CREATE TABLE event_types (
          event_type_id INTEGER PRIMARY KEY,
          event_type TEXT
        );

        CREATE TABLE events (
          event_id INTEGER PRIMARY KEY,
          nga_west2_eqid INTEGER,
          hypocenter_longitude REAL,
          hypocenter_latitude REAL,
          event_country TEXT,
          event_subdivision TEXT,
          event_name TEXT,
          comcat_id TEXT,
          event_type_id INTEGER,
          datetime TEXT,
          time_zone TEXT,
          hypocenter_reported_depth REAL,
          hypocenter_reported_depth_datum TEXT,
          epicenter_elevation REAL,
          hypocenter_depth REAL,
          magnitude REAL,
          magnitude_type TEXT,
          magnitude_uncertainty_kagan_model REAL,
          magnitude_uncertainty_statistical REAL,
          magnitude_sample_size INTEGER,
          magnitude_uncertainty_study_class TEXT,
          strike REAL,
          dip REAL,
          rake REAL,
          mechanism_based_on_rake TEXT,
          slip_rate REAL,
          seismic_moment REAL,
          p_plunge REAL,
          p_trend REAL,
          t_plunge REAL,
          t_trend REAL,
          coseismic_surface_rupture INTEGER,
          basis_for_surface_rupture TEXT,
          extensional_regime INTEGER,
          extensional_regime_name TEXT,
          synchronous_event INTEGER,
          event_type TEXT
        );

        CREATE TABLE finite_faults (
          finite_fault_id INTEGER PRIMARY KEY,
          event_id INTEGER,
          ztor REAL,
          fault_length REAL,
          fault_width REAL,
          fault_area REAL,
          ffm_model TEXT,
          ffm_complexity INTEGER,
          finite_fault_citation_id INTEGER,
          finite_fault_kinematic_parameter_id INTEGER
        );

        CREATE TABLE finite_fault_kinematic_parameters (
          finite_fault_kinematic_parameter_id INTEGER PRIMARY KEY,
          finite_fault_id INTEGER,
          event_id INTEGER,
          average_fault_displacement REAL,
          rise_time REAL,
          average_slip_velocity REAL,
          preferred_rupture_velocity REAL,
          average_vr_vs REAL,
          percent_moment_release REAL,
          existence_of_shallow_asperity INTEGER,
          depth_to_shallowest_asperity REAL
        );

        CREATE TABLE finite_fault_segments (
          event_id INTEGER,
          finite_fault_id INTEGER,
          segment_index INTEGER,
          seg_sub_rupture_number INTEGER,
          ulc_latitude REAL,
          ulc_longitude REAL,
          ulc_depth REAL,
          seg_length REAL,
          seg_width REAL,
          seg_area REAL,
          seg_strike REAL,
          seg_dip REAL,
          seg_rake REAL,
          PRIMARY KEY (event_id, finite_fault_id, segment_index)
        );

        CREATE TABLE networks (
          network_id INTEGER PRIMARY KEY,
          network_code TEXT,
          network_name TEXT,
          network_type TEXT,
          start_date TEXT,
          end_date TEXT,
          operation_org TEXT,
          network_citation_id INTEGER
        );

        CREATE TABLE stations (
          station_id INTEGER PRIMARY KEY,
          nga_west2_ssn INTEGER,
          site_id INTEGER,
          network_id INTEGER,
          station_name TEXT,
          station_latitude REAL,
          station_longitude REAL,
          station_code TEXT,
          housing TEXT,
          cosmos_station_type INTEGER,
          sensor_depth REAL,
          installation_date TEXT,
          removal_date TEXT
        );

        CREATE TABLE sites (
          site_id INTEGER PRIMARY KEY,
          site_longitude REAL,
          site_latitude REAL,
          site_elevation REAL,
          site_country TEXT,
          site_subdivision TEXT,
          site_name TEXT,
          vs30 REAL,
          vs30_lnstd REAL,
          vs30_code_id INTEGER,
          vs30_ref TEXT,
          geological_unit TEXT,
          geological_citation_id INTEGER,
          slope_gradient REAL,
          slope_resolution REAL,
          terrain_class TEXT,
          terrain_citation_id INTEGER,
          z1p0_preferred REAL,
          z1p0_preferred_lnstd REAL,
          z1p0_code_id INTEGER,
          z2p5_preferred REAL,
          z2p5_preferred_lnstd REAL,
          z2p5_code_id INTEGER,
          basin_geomorphic_category TEXT,
          basin_geospatial_category TEXT,
          gmx_c2 TEXT,
          gmx_c3 TEXT,
          rsbe REAL,
          rcebe REAL,
          division TEXT,
          province TEXT,
          section TEXT,
          geological_unit_cgs TEXT
        );

        CREATE TABLE basin_depth_estimates (
          site_id INTEGER,
          model_name TEXT,
          z1p0 REAL,
          z1p0_lnstd REAL,
          z2p5 REAL,
          z2p5_lnstd REAL,
          PRIMARY KEY (site_id, model_name)
        );

        CREATE TABLE motions (
          motion_id INTEGER PRIMARY KEY,
          event_id INTEGER,
          station_id INTEGER,
          path_id INTEGER,
          user_id INTEGER,
          public_motion INTEGER,
          nga_west2_rsn INTEGER,
          nga_west2_eqid INTEGER,
          nga_west2_ssn INTEGER,
          nyquist_frequency REAL
        );

        CREATE TABLE paths (
          path_id INTEGER PRIMARY KEY,
          motion_id INTEGER,
          epicentral_distance REAL,
          hypocentral_distance REAL,
          rjb REAL,
          rrup REAL,
          rx REAL,
          ry REAL,
          ry0 REAL,
          ravg REAL,
          rebe REAL,
          rsbe1 REAL,
          closest_point_latitude REAL,
          closest_point_longitude REAL,
          closest_point_depth REAL,
          hanging_wall INTEGER,
          source_to_site_azimuth REAL,
          x REAL,
          theta_d REAL,
          y REAL,
          phi_d REAL,
          ssga_strike_slip REAL,
          ssga_dip_slip REAL,
          s REAL,
          d REAL,
          ctildepr REAL,
          t REAL,
          u REAL
        );

        CREATE TABLE time_series_metadata (
          time_series_metadata_id INTEGER PRIMARY KEY,
          motion_id INTEGER,
          time_series_citation_id INTEGER,
          instrument_type INTEGER,
          colocated_instrument INTEGER,
          instrument_natural_frequency REAL,
          sampling_rate REAL,
          late_p_trigger INTEGER,
          late_s_trigger INTEGER,
          multiple_wave_train INTEGER,
          processing_type INTEGER,
          filter_type INTEGER,
          order_hpass INTEGER,
          order_lpass INTEGER,
          azimuth_h1 REAL,
          azimuth_h2 REAL,
          hpass_applied_h1 INTEGER,
          hpass_applied_h2 INTEGER,
          hpass_applied_v INTEGER,
          hpass_fc_h1 REAL,
          hpass_fc_h2 REAL,
          hpass_fc_v REAL,
          lpass_applied_h1 INTEGER,
          lpass_applied_h2 INTEGER,
          lpass_applied_v INTEGER,
          lpass_fc_h1 REAL,
          lpass_fc_h2 REAL,
          lpass_fc_v REAL,
          usable_frequency_factor REAL
        );

        CREATE TABLE intensity_measures (
          motion_id INTEGER,
          component TEXT,
          pga REAL,
          pgv REAL,
          cav REAL,
          cav5 REAL,
          ia REAL,
          ia_time_05 REAL,
          ia_time_10 REAL,
          ia_time_15 REAL,
          ia_time_20 REAL,
          ia_time_25 REAL,
          ia_time_30 REAL,
          ia_time_35 REAL,
          ia_time_40 REAL,
          ia_time_45 REAL,
          ia_time_50 REAL,
          ia_time_55 REAL,
          ia_time_60 REAL,
          ia_time_65 REAL,
          ia_time_70 REAL,
          ia_time_75 REAL,
          ia_time_80 REAL,
          ia_time_85 REAL,
          ia_time_90 REAL,
          ia_time_95 REAL,
          PRIMARY KEY (motion_id, component)
        );

        CREATE TABLE spectral_axes (
          spectrum_type TEXT NOT NULL,
          component TEXT NOT NULL,
          ordinate_index INTEGER NOT NULL,
          axis_value REAL NOT NULL,
          axis_unit TEXT NOT NULL,
          source_field TEXT NOT NULL,
          PRIMARY KEY (spectrum_type, component, ordinate_index)
        );

        CREATE TABLE response_spectra (
          motion_id INTEGER,
          component TEXT,
          damping REAL,
          psa_json TEXT NOT NULL,
          PRIMARY KEY (motion_id, component)
        );

        CREATE TABLE effective_amplitude_spectra (
          motion_id INTEGER PRIMARY KEY,
          konno_omachi_points INTEGER,
          smoothing_bandwidth REAL,
          window_width REAL,
          eas_json TEXT NOT NULL
        );

        CREATE TABLE c1c2_classifications (
          event_id INTEGER,
          rcutoff_km REAL,
          rrup REAL,
          rjb REAL,
          rx REAL,
          delta_time REAL,
          time_window REAL,
          class TEXT,
          cluster_num INTEGER,
          magnitude REAL,
          PRIMARY KEY (event_id, rcutoff_km)
        );
        """
    )


def read_csv_dicts(path: Path):
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as fh:
        yield from csv.DictReader(fh)


def insert_rows(conn: sqlite3.Connection, table: str, columns: list[str], rows) -> int:
    sql = f"INSERT OR REPLACE INTO {q(table)} ({','.join(q(c) for c in columns)}) VALUES ({','.join('?' for _ in columns)})"
    count = 0
    batch = []
    for row in rows:
        batch.append([row.get(c) for c in columns])
        if len(batch) >= BATCH_SIZE:
            conn.executemany(sql, batch)
            count += len(batch)
            batch.clear()
    if batch:
        conn.executemany(sql, batch)
        count += len(batch)
    return count


def load_documentation(conn: sqlite3.Connection) -> None:
    wb = load_workbook(DOC_XLSX, read_only=True, data_only=True)
    for sheet in [
        "PSA Field Descriptions",
        "EAS Field Descriptions",
        "Source Metadata",
        "Station Metadata",
    ]:
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        next(rows, None)
        empty_run = 0
        batch = []
        for row in rows:
            col, field, dtype, desc = (list(row) + [None] * 4)[:4]
            if field is None:
                empty_run += 1
                if empty_run > 25:
                    break
                continue
            empty_run = 0
            batch.append((sheet, str(col) if col is not None else None, str(field), str(dtype) if dtype else None, str(desc) if desc else None))
        conn.executemany(
            "INSERT OR REPLACE INTO field_catalog VALUES (?,?,?,?,?)",
            batch,
        )

    if "Citations" in wb.sheetnames:
        ws = wb["Citations"]
        rows = ws.iter_rows(values_only=True)
        next(rows, None)
        batch = []
        for row in rows:
            vals = list(row)
            if not vals or vals[0] is None:
                continue
            try:
                citation_id = int(vals[0])
            except (TypeError, ValueError):
                continue
            batch.append((citation_id, vals[1] if len(vals) > 1 else None, vals[2] if len(vals) > 2 else None))
        conn.executemany("INSERT OR REPLACE INTO citations VALUES (?,?,?)", batch)

    for sheet in ["Housing", "COSMOS Station Type", "VS30 Codes", "Z Codes"]:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(x) if x is not None else f"col_{i}" for i, x in enumerate(rows[0])]
        batch = []
        for row in rows[1:]:
            if not row or row[0] is None:
                continue
            raw = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            code = str(row[0])
            label = str(row[1]) if len(row) > 1 and row[1] is not None else None
            desc = str(row[2]) if len(row) > 2 and row[2] is not None else label
            batch.append((sheet, code, label, desc, json.dumps(raw, ensure_ascii=False, default=str)))
        conn.executemany("INSERT OR REPLACE INTO code_definitions VALUES (?,?,?,?,?)", batch)


def load_release_files(conn: sqlite3.Connection) -> None:
    files = [
        (SOURCE_CSV, "source_metadata", None),
        (STATION_CSV, "station_metadata", None),
        (DOC_XLSX, "documentation", None),
        (C1C2_XLSX, "supplement", None),
        (EAS_CSV, "ground_motion_eas", "EAS"),
    ] + [(path, "ground_motion_psa", comp) for comp, path in COMPONENT_FILES.items()]
    batch = []
    for path, role, comp in files:
        rows = None
        cols = None
        if path.suffix.lower() == ".csv":
            header = csv_header(path)
            cols = len(header)
            with path.open("rb") as fh:
                rows = sum(1 for _ in fh) - 1
        batch.append((path.name, role, comp, rows, cols, sha256_file(path)))
    conn.executemany("INSERT OR REPLACE INTO release_files VALUES (?,?,?,?,?,?)", batch)


def load_source(conn: sqlite3.Connection) -> None:
    event_cols_src = [
        "event_id",
        "NGA_West2_EQID",
        "hypocenter_longitude",
        "hypocenter_latitude",
        "event_country",
        "event_subdivision",
        "event_name",
        "comcat_id",
        "event_type_id",
        "datetime",
        "time_zone",
        "hypocenter_reported_depth",
        "hypocenter_reported_depth_datum",
        "epicenter_elevation",
        "hypocenter_depth",
        "magnitude",
        "magnitude_type",
        "magnitude_uncertainty_Kagan_model",
        "magnitude_uncertainty_statistical",
        "magnitude_sample_size",
        "magnitude_uncertainty_study_class",
        "strike",
        "dip",
        "rake",
        "mechanism_based_on_Rake",
        "slip_rate",
        "seismic_moment",
        "p_plunge",
        "p_trend",
        "t_plunge",
        "t_trend",
        "coseismic_surface_rupture",
        "basis_for_surface_rupture",
        "extensional_regime",
        "extensional_regime_name",
        "synchronous_event",
        "event_type",
    ]
    event_cols_db = [
        "event_id",
        "nga_west2_eqid",
        "hypocenter_longitude",
        "hypocenter_latitude",
        "event_country",
        "event_subdivision",
        "event_name",
        "comcat_id",
        "event_type_id",
        "datetime",
        "time_zone",
        "hypocenter_reported_depth",
        "hypocenter_reported_depth_datum",
        "epicenter_elevation",
        "hypocenter_depth",
        "magnitude",
        "magnitude_type",
        "magnitude_uncertainty_kagan_model",
        "magnitude_uncertainty_statistical",
        "magnitude_sample_size",
        "magnitude_uncertainty_study_class",
        "strike",
        "dip",
        "rake",
        "mechanism_based_on_rake",
        "slip_rate",
        "seismic_moment",
        "p_plunge",
        "p_trend",
        "t_plunge",
        "t_trend",
        "coseismic_surface_rupture",
        "basis_for_surface_rupture",
        "extensional_regime",
        "extensional_regime_name",
        "synchronous_event",
        "event_type",
    ]
    finite_cols = [
        "finite_fault_id",
        "event_id",
        "ztor",
        "fault_length",
        "fault_width",
        "fault_area",
        "ffm_model",
        "ffm_complexity",
        "finite_fault_citation_id",
        "finite_fault_kinematic_parameter_id",
    ]
    kin_cols = [
        "finite_fault_kinematic_parameter_id",
        "finite_fault_id",
        "event_id",
        "average_fault_displacement",
        "rise_time",
        "average_slip_velocity",
        "preferred_rupture_velocity",
        "average_vr_vs",
        "percent_moment_release",
        "existence_of_shallow_asperity",
        "depth_to_shallowest_asperity",
    ]
    ev_batch, type_batch, ff_batch, kin_batch, seg_batch = [], [], [], [], []
    for row in read_csv_dicts(SOURCE_CSV):
        clean = {k: scalar(v) for k, v in row.items()}
        ev_batch.append({db: clean.get(src) for src, db in zip(event_cols_src, event_cols_db)})
        if clean.get("event_type_id") is not None:
            type_batch.append((clean.get("event_type_id"), clean.get("event_type")))
        if clean.get("finite_fault_id") is not None:
            ff_batch.append({c: clean.get(c) for c in finite_cols})
        if clean.get("finite_fault_kinematic_parameter_id") is not None:
            item = {c: clean.get(c) for c in kin_cols}
            if has_real_value(item.values()):
                kin_batch.append(item)
        for i in range(1, 17):
            vals = {
                "event_id": clean.get("event_id"),
                "finite_fault_id": clean.get("finite_fault_id"),
                "segment_index": i,
                "seg_sub_rupture_number": clean.get(f"seg_sub_rupture_number_{i}"),
                "ulc_latitude": clean.get(f"ULC_latitude_{i}"),
                "ulc_longitude": clean.get(f"ULC_longitude_{i}"),
                "ulc_depth": clean.get(f"ULC_depth_{i}"),
                "seg_length": clean.get(f"seg_length_{i}"),
                "seg_width": clean.get(f"seg_width_{i}"),
                "seg_area": clean.get(f"seg_area_{i}"),
                "seg_strike": clean.get(f"seg_strike_{i}"),
                "seg_dip": clean.get(f"seg_dip_{i}"),
                "seg_rake": clean.get(f"seg_rake_{i}"),
            }
            if has_real_value(vals.values()):
                seg_batch.append(vals)
    insert_rows(conn, "events", event_cols_db, ev_batch)
    conn.executemany("INSERT OR REPLACE INTO event_types VALUES (?,?)", type_batch)
    insert_rows(conn, "finite_faults", finite_cols, ff_batch)
    insert_rows(conn, "finite_fault_kinematic_parameters", kin_cols, kin_batch)
    insert_rows(conn, "finite_fault_segments", list(seg_batch[0].keys()), seg_batch)


def load_stations(conn: sqlite3.Connection) -> None:
    network_cols = [
        "network_id",
        "network_code",
        "network_name",
        "network_type",
        "start_date",
        "end_date",
        "operation_org",
        "network_citation_id",
    ]
    station_cols_src = [
        "station_id",
        "NGA_West2_SSN",
        "site_id",
        "network_id",
        "station_name",
        "station_latitude",
        "station_longitude",
        "station_code",
        "housing",
        "cosmos_station_type",
        "sensor_depth",
        "installation_date",
        "removal_date",
    ]
    station_cols_db = [
        "station_id",
        "nga_west2_ssn",
        "site_id",
        "network_id",
        "station_name",
        "station_latitude",
        "station_longitude",
        "station_code",
        "housing",
        "cosmos_station_type",
        "sensor_depth",
        "installation_date",
        "removal_date",
    ]
    site_cols = [
        "site_id",
        "site_longitude",
        "site_latitude",
        "site_elevation",
        "site_country",
        "site_subdivision",
        "site_name",
        "vs30",
        "vs30_lnstd",
        "vs30_code_id",
        "vs30_ref",
        "geological_unit",
        "geological_citation_id",
        "slope_gradient",
        "slope_resolution",
        "terrain_class",
        "terrain_citation_id",
        "z1p0_preferred",
        "z1p0_preferred_lnstd",
        "z1p0_code_id",
        "z2p5_preferred",
        "z2p5_preferred_lnstd",
        "z2p5_code_id",
        "basin_geomorphic_category",
        "basin_geospatial_category",
        "gmx_c2",
        "gmx_c3",
        "rsbe",
        "rcebe",
        "division",
        "province",
        "section",
        "geological_unit_cgs",
    ]
    src_case_map = {
        "division": "DIVISION",
        "province": "PROVINCE",
        "section": "SECTION",
    }
    models = [
        "measured",
        "extrapolated",
        "CAGeo",
        "CVMS4",
        "CVMS4.26",
        "CVMS4.26.M01",
        "CVMH15.1",
        "SFCVM21.1",
        "GreatValley",
        "WFCVM",
        "USGSNCM",
        "NCREE",
        "NIED",
        "JSHIS",
        "NZGeo",
        "NZVM",
    ]
    networks, stations, sites, basin = {}, {}, {}, {}
    for row in read_csv_dicts(STATION_CSV):
        clean = {k: scalar(v) for k, v in row.items()}
        if clean.get("network_id") is not None:
            networks[clean["network_id"]] = {c: clean.get(c) for c in network_cols}
        stations[clean["station_id"]] = {db: clean.get(src) for src, db in zip(station_cols_src, station_cols_db)}
        site = {}
        for c in site_cols:
            src = src_case_map.get(c, c)
            site[c] = clean.get(src)
        sites[clean["site_id"]] = site
        for model in models:
            vals = {
                "site_id": clean.get("site_id"),
                "model_name": model,
                "z1p0": clean.get(f"z1p0_{model}"),
                "z1p0_lnstd": clean.get(f"z1p0_lnstd_{model}"),
                "z2p5": clean.get(f"z2p5_{model}"),
                "z2p5_lnstd": clean.get(f"z2p5_lnstd_{model}"),
            }
            if has_real_value([vals["z1p0"], vals["z1p0_lnstd"], vals["z2p5"], vals["z2p5_lnstd"]]):
                basin[(vals["site_id"], model)] = vals
    insert_rows(conn, "networks", network_cols, networks.values())
    insert_rows(conn, "stations", station_cols_db, stations.values())
    insert_rows(conn, "sites", site_cols, sites.values())
    insert_rows(conn, "basin_depth_estimates", ["site_id", "model_name", "z1p0", "z1p0_lnstd", "z2p5", "z2p5_lnstd"], basin.values())


def load_motion_core(conn: sqlite3.Connection) -> None:
    path = COMPONENT_FILES[MOTION_BASE_COMPONENT]
    path_cols = [
        "path_id",
        "motion_id",
        "epicentral_distance",
        "hypocentral_distance",
        "rjb",
        "rrup",
        "rx",
        "ry",
        "ry0",
        "ravg",
        "rebe",
        "rsbe1",
        "closest_point_latitude",
        "closest_point_longitude",
        "closest_point_depth",
        "hanging_wall",
        "source_to_site_azimuth",
        "x",
        "theta_d",
        "y",
        "phi_d",
        "ssga_strike_slip",
        "ssga_dip_slip",
        "s",
        "d",
        "ctildepr",
        "t",
        "u",
    ]
    path_src = {c: c for c in path_cols}
    path_src.update({"x": "X", "theta_d": "theta_D", "y": "Y", "phi_d": "phi_D", "t": "T", "u": "U"})
    motion_cols = [
        "motion_id",
        "event_id",
        "station_id",
        "path_id",
        "user_id",
        "public_motion",
        "nga_west2_rsn",
        "nga_west2_eqid",
        "nga_west2_ssn",
        "nyquist_frequency",
    ]
    ts_cols = [
        "time_series_metadata_id",
        "motion_id",
        "time_series_citation_id",
        "instrument_type",
        "colocated_instrument",
        "instrument_natural_frequency",
        "sampling_rate",
        "late_p_trigger",
        "late_s_trigger",
        "multiple_wave_train",
        "processing_type",
        "filter_type",
        "order_hpass",
        "order_lpass",
        "azimuth_h1",
        "azimuth_h2",
        "hpass_applied_h1",
        "hpass_applied_h2",
        "hpass_applied_v",
        "hpass_fc_h1",
        "hpass_fc_h2",
        "hpass_fc_v",
        "lpass_applied_h1",
        "lpass_applied_h2",
        "lpass_applied_v",
        "lpass_fc_h1",
        "lpass_fc_h2",
        "lpass_fc_v",
        "usable_frequency_factor",
    ]
    motion_batch, path_batch, ts_batch = [], [], []
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as fh:
        for row in csv.DictReader(fh):
            clean = {k: scalar(v) for k, v in row.items()}
            if clean.get("motion_id") in (None, -999):
                continue
            motion_batch.append(
                {
                    "motion_id": clean.get("motion_id"),
                    "event_id": clean.get("event_id"),
                    "station_id": clean.get("station_id"),
                    "path_id": clean.get("path_id"),
                    "user_id": clean.get("user_id"),
                    "public_motion": clean.get("public_motion"),
                    "nga_west2_rsn": clean.get("NGA_West2_RSN"),
                    "nga_west2_eqid": clean.get("NGA_West2_EQID"),
                    "nga_west2_ssn": clean.get("NGA_West2_SSN"),
                    "nyquist_frequency": clean.get("nyquist_frequency"),
                }
            )
            path_batch.append({c: clean.get(path_src[c]) for c in path_cols})
            ts_batch.append({c: clean.get(c) for c in ts_cols})
            if len(motion_batch) >= BATCH_SIZE:
                insert_rows(conn, "motions", motion_cols, motion_batch)
                insert_rows(conn, "paths", path_cols, path_batch)
                insert_rows(conn, "time_series_metadata", ts_cols, ts_batch)
                motion_batch.clear()
                path_batch.clear()
                ts_batch.clear()
    insert_rows(conn, "motions", motion_cols, motion_batch)
    insert_rows(conn, "paths", path_cols, path_batch)
    insert_rows(conn, "time_series_metadata", ts_cols, ts_batch)


def json_array(row: dict[str, str], columns: list[str]) -> str:
    return json.dumps([float_or_none(row.get(c)) for c in columns], separators=(",", ":"))


def load_psa_components(conn: sqlite3.Connection) -> None:
    im_cols = [
        "motion_id",
        "component",
        "pga",
        "pgv",
        "cav",
        "cav5",
        "ia",
        "ia_time_05",
        "ia_time_10",
        "ia_time_15",
        "ia_time_20",
        "ia_time_25",
        "ia_time_30",
        "ia_time_35",
        "ia_time_40",
        "ia_time_45",
        "ia_time_50",
        "ia_time_55",
        "ia_time_60",
        "ia_time_65",
        "ia_time_70",
        "ia_time_75",
        "ia_time_80",
        "ia_time_85",
        "ia_time_90",
        "ia_time_95",
    ]
    for component, path in COMPONENT_FILES.items():
        header = csv_header(path)
        psa_cols = []
        periods = []
        for col in header:
            match = re.match(r"^psa_[^(]+\(([^)]+)s\)$", col)
            if match:
                psa_cols.append(col)
                periods.append(parse_axis_value(match.group(1)))
        conn.executemany(
            "INSERT OR REPLACE INTO spectral_axes VALUES (?,?,?,?,?,?)",
            [("PSA", component, i, period, "s", col) for i, (period, col) in enumerate(zip(periods, psa_cols), start=1)],
        )

        im_batch = []
        spec_batch = []
        with path.open(newline="", encoding="utf-8-sig", errors="replace") as fh:
            for row in csv.DictReader(fh):
                motion_id = scalar(row.get("motion_id"))
                if motion_id in (None, -999):
                    continue
                lower_cols = {c.lower(): c for c in row.keys()}
                comp_suffix = component.lower()
                im = {
                    "motion_id": motion_id,
                    "component": component,
                    "pga": scalar(row.get(lower_cols.get(f"pga_{comp_suffix}", ""))),
                    "pgv": scalar(row.get(lower_cols.get(f"pgv_{comp_suffix}", ""))),
                    "cav": scalar(row.get(lower_cols.get(f"cav_{comp_suffix}", ""))),
                    "cav5": scalar(row.get(lower_cols.get(f"cav5_{comp_suffix}", ""))),
                    "ia": scalar(row.get(lower_cols.get(f"ia_{comp_suffix}", ""))),
                }
                for pct in range(5, 100, 5):
                    key = f"ia_time_{pct:02d}_{comp_suffix}"
                    im[f"ia_time_{pct:02d}"] = scalar(row.get(lower_cols.get(key, "")))
                im_batch.append(im)
                spec_batch.append(
                    (
                        motion_id,
                        component,
                        scalar(row.get("damping")),
                        json_array(row, psa_cols),
                    )
                )
                if len(im_batch) >= BATCH_SIZE:
                    insert_rows(conn, "intensity_measures", im_cols, im_batch)
                    conn.executemany("INSERT OR REPLACE INTO response_spectra VALUES (?,?,?,?)", spec_batch)
                    im_batch.clear()
                    spec_batch.clear()
        insert_rows(conn, "intensity_measures", im_cols, im_batch)
        conn.executemany("INSERT OR REPLACE INTO response_spectra VALUES (?,?,?,?)", spec_batch)
        conn.commit()
        print(f"Loaded {component}")


def load_eas(conn: sqlite3.Connection) -> None:
    header = csv_header(EAS_CSV)
    eas_cols = []
    freqs = []
    for col in header:
        match = re.match(r"^eas\(([^)]+)Hz\)$", col)
        if match:
            eas_cols.append(col)
            freqs.append(parse_axis_value(match.group(1)))
    conn.executemany(
        "INSERT OR REPLACE INTO spectral_axes VALUES (?,?,?,?,?,?)",
        [("EAS", "EAS", i, freq, "Hz", col) for i, (freq, col) in enumerate(zip(freqs, eas_cols), start=1)],
    )
    batch = []
    with EAS_CSV.open(newline="", encoding="utf-8-sig", errors="replace") as fh:
        for row in csv.DictReader(fh):
            motion_id = scalar(row.get("motion_id"))
            if motion_id in (None, -999):
                continue
            batch.append(
                (
                    motion_id,
                    scalar(row.get("konno_omachi_points")),
                    scalar(row.get("smoothing_bandwidth")),
                    scalar(row.get("window_width")),
                    json_array(row, eas_cols),
                )
            )
            if len(batch) >= BATCH_SIZE:
                conn.executemany("INSERT OR REPLACE INTO effective_amplitude_spectra VALUES (?,?,?,?,?)", batch)
                batch.clear()
    conn.executemany("INSERT OR REPLACE INTO effective_amplitude_spectra VALUES (?,?,?,?,?)", batch)


def load_c1c2(conn: sqlite3.Connection) -> None:
    wb = load_workbook(C1C2_XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 5:
        return
    batch = []
    for block_start in range(0, ws.max_column, 11):
        cutoff = rows[0][block_start + 1] if block_start + 1 < len(rows[0]) else None
        if cutoff is None:
            continue
        for row in rows[4:]:
            if block_start + 9 >= len(row) or row[block_start] is None:
                continue
            batch.append(
                (
                    scalar(row[block_start]),
                    scalar(cutoff),
                    scalar(row[block_start + 1]),
                    scalar(row[block_start + 2]),
                    scalar(row[block_start + 3]),
                    scalar(row[block_start + 5]),
                    scalar(row[block_start + 6]),
                    scalar(row[block_start + 7]),
                    scalar(row[block_start + 8]),
                    scalar(row[block_start + 9]),
                )
            )
    conn.executemany("INSERT OR REPLACE INTO c1c2_classifications VALUES (?,?,?,?,?,?,?,?,?,?)", batch)


def create_indexes(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE INDEX IF NOT EXISTS idx_motions_event ON motions(event_id);
        CREATE INDEX IF NOT EXISTS idx_motions_station ON motions(station_id);
        CREATE INDEX IF NOT EXISTS idx_paths_motion ON paths(motion_id);
        CREATE INDEX IF NOT EXISTS idx_tsm_motion ON time_series_metadata(motion_id);
        CREATE INDEX IF NOT EXISTS idx_im_component ON intensity_measures(component);
        CREATE INDEX IF NOT EXISTS idx_psa_component ON response_spectra(component);
        CREATE INDEX IF NOT EXISTS idx_events_region ON events(event_country, event_subdivision);
        CREATE INDEX IF NOT EXISTS idx_sites_region ON sites(site_country, site_subdivision);
        CREATE INDEX IF NOT EXISTS idx_c1c2_class ON c1c2_classifications(class);
        ANALYZE;
        """
    )


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    if OUT.exists():
        OUT.unlink()
    conn = sqlite3.connect(OUT)
    try:
        create_schema(conn)
        load_documentation(conn)
        load_release_files(conn)
        conn.commit()
        print("Loaded documentation")
        load_source(conn)
        conn.commit()
        print("Loaded source metadata")
        load_stations(conn)
        conn.commit()
        print("Loaded station metadata")
        load_c1c2(conn)
        conn.commit()
        print("Loaded C1/C2 supplementary classifications")
        load_motion_core(conn)
        conn.commit()
        print("Loaded motion/path/time-series core")
        load_psa_components(conn)
        conn.commit()
        load_eas(conn)
        conn.commit()
        print("Loaded EAS")
        create_indexes(conn)
        conn.commit()
    finally:
        conn.close()
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
